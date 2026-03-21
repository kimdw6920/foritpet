from django.contrib import admin
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.urls import path
from django.utils import timezone
from django.contrib import messages
from openpyxl import load_workbook

from .models import Product, Shelter, Donation


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'price')


@admin.register(Shelter)
class ShelterAdmin(admin.ModelAdmin):
    list_display = ('name', 'region', 'address', 'phone')
    list_filter = ('region',)


@staff_member_required
def upload_tracking_view(request):
    """송장 엑셀 업로드: 후원ID, 송장번호, 택배사 컬럼으로 Donation 매칭 후 저장."""
    if request.method != 'POST':
        return render(request, 'admin/shelter/donation/upload_tracking.html')

    f = request.FILES.get('excel_file')
    if not f or not f.name.lower().endswith('.xlsx'):
        messages.error(request, '엑셀 파일(.xlsx)을 선택해 주세요.')
        return redirect('admin:shelter_donation_upload_tracking')

    updated = 0
    errors = []
    try:
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
    except Exception as e:
        messages.error(request, f'엑셀 읽기 실패: {e}')
        return redirect('admin:shelter_donation_upload_tracking')

    if not rows:
        messages.warning(request, '데이터가 없습니다.')
        return redirect('admin:shelter_donation_upload_tracking')

    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    idx_id, idx_num, idx_carrier = 0, 1, 2
    for i, h in enumerate(header[:5]):
        if '후원' in h and ('id' in h.lower() or h == '후원ID'):
            idx_id = i
        if '송장' in h or '번호' in h:
            idx_num = i
        if '택배' in h or 'carrier' in h.lower():
            idx_carrier = i

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or len(row) <= max(idx_id, idx_num, idx_carrier):
            continue
        try:
            raw_id = row[idx_id]
            donation_id = int(raw_id) if raw_id is not None else None
        except (TypeError, ValueError):
            errors.append(f'{row_num}행: 후원ID 숫자가 아님')
            continue
        tracking_number = str(row[idx_num]).strip() if row[idx_num] is not None else ''
        tracking_carrier = str(row[idx_carrier]).strip() if row[idx_carrier] is not None else ''
        if not donation_id or not tracking_number:
            continue
        try:
            donation = Donation.objects.get(pk=donation_id)
            donation.tracking_number = tracking_number
            donation.tracking_carrier = tracking_carrier
            donation.shipped_at = donation.shipped_at or timezone.now()
            donation.save()
            updated += 1
        except Donation.DoesNotExist:
            errors.append(f'{row_num}행: 후원ID {donation_id} 없음')
        except Exception as e:
            errors.append(f'{row_num}행: {e}')

    if updated:
        messages.success(request, f'송장 {updated}건 반영되었습니다.')
    for err in errors[:10]:
        messages.warning(request, err)
    if len(errors) > 10:
        messages.warning(request, f'외 오류 {len(errors) - 10}건 생략.')
    return redirect('admin:shelter_donation_upload_tracking')


@admin.register(Donation)
class DonationAdmin(admin.ModelAdmin):
    list_display = ('id', 'user', 'shelter', 'product', 'amount', 'tracking_carrier', 'tracking_number', 'created_at')
    list_filter = ('shelter',)

    def get_urls(self):
        urls = super().get_urls()
        extra = [
            path('upload-tracking/', upload_tracking_view, name='shelter_donation_upload_tracking'),
        ]
        return extra + urls